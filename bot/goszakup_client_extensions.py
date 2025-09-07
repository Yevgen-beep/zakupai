"""
Расширения для GoszakupClient: экспорт, мониторинг, батчинг и телеграм интеграции
"""

import asyncio
import csv
import json
import logging
from collections.abc import Callable
from datetime import datetime
from io import StringIO
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from goszakup_client_v3 import (
    ContractFiltersInput,
    ContractResult,
    GoszakupClient,
    LotResult,
    LotsFiltersInput,
    LotStatus,
    SubjectResult,
    TradeMethod,
)

logger = logging.getLogger(__name__)


class ExportFormat:
    """Форматы экспорта"""

    JSON = "json"
    CSV = "csv"
    EXCEL = "xlsx"
    TXT = "txt"


class ExportMixin:
    """Миксин для добавления функций экспорта"""

    async def export_search_results(
        self,
        results: list[LotResult | ContractResult | SubjectResult],
        format_type: str = ExportFormat.JSON,
        filename: str | None = None,
        include_metadata: bool = True,
    ) -> str | bytes:
        """
        Экспорт результатов поиска в различные форматы

        Args:
            results: Результаты поиска для экспорта
            format_type: Формат экспорта (json, csv, xlsx, txt)
            filename: Имя файла (опционально)
            include_metadata: Включить метаданные

        Returns:
            Данные в указанном формате
        """
        if not results:
            return ""

        # Определение типа результатов (удалена неиспользуемая переменная)

        if format_type == ExportFormat.JSON:
            return await self._export_to_json(results, include_metadata)
        elif format_type == ExportFormat.CSV:
            return await self._export_to_csv(results)
        elif format_type == ExportFormat.EXCEL and EXCEL_AVAILABLE:
            return await self._export_to_excel(results, filename)
        elif format_type == ExportFormat.TXT:
            return await self._export_to_txt(results)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")

    async def _export_to_json(
        self, results: list[Any], include_metadata: bool = True
    ) -> str:
        """Экспорт в JSON"""
        data = {
            "results": [
                result.to_dict() if hasattr(result, "to_dict") else result
                for result in results
            ],
            "count": len(results),
        }

        if include_metadata:
            data["metadata"] = {
                "export_time": datetime.now().isoformat(),
                "result_type": type(results[0]).__name__ if results else "Unknown",
                "format": "JSON",
            }

        return json.dumps(data, ensure_ascii=False, indent=2)

    async def _export_to_csv(self, results: list[Any]) -> str:
        """Экспорт в CSV"""
        if not results:
            return ""

        output = StringIO()

        # Получение данных первой записи для определения полей
        first_item = (
            results[0].to_dict() if hasattr(results[0], "to_dict") else results[0]
        )

        if isinstance(first_item, dict):
            fieldnames = list(first_item.keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames, lineterminator="\n")
            writer.writeheader()

            for result in results:
                row_data = result.to_dict() if hasattr(result, "to_dict") else result
                # Преобразование сложных объектов в строки
                for key, value in row_data.items():
                    if isinstance(value, dict | list):
                        row_data[key] = json.dumps(value, ensure_ascii=False)
                writer.writerow(row_data)

        return output.getvalue()

    async def _export_to_excel(
        self, results: list[Any], filename: str | None = None
    ) -> bytes:
        """Экспорт в Excel"""
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl не установлен. Установите: pip install openpyxl"
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Результаты поиска"

        if not results:
            return b""

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        # Получение данных для заголовков
        first_item = (
            results[0].to_dict() if hasattr(results[0], "to_dict") else results[0]
        )
        headers = list(first_item.keys())

        # Запись заголовков
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Запись данных
        for row, result in enumerate(results, 2):
            row_data = result.to_dict() if hasattr(result, "to_dict") else result
            for col, (_key, value) in enumerate(row_data.items(), 1):
                # Преобразование сложных типов
                if isinstance(value, dict | list):
                    value = json.dumps(value, ensure_ascii=False)
                elif value is None:
                    value = ""
                worksheet.cell(row=row, column=col, value=str(value))

        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Сохранение в байты
        from io import BytesIO

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def _export_to_txt(self, results: list[Any]) -> str:
        """Экспорт в текстовый формат"""
        lines = []
        lines.append(f"Результаты поиска ({len(results)} записей)")
        lines.append("=" * 50)
        lines.append("")

        for i, result in enumerate(results, 1):
            lines.append(f"Запись {i}:")
            lines.append("-" * 20)

            data = result.to_dict() if hasattr(result, "to_dict") else result
            for key, value in data.items():
                if isinstance(value, dict | list):
                    value = json.dumps(value, ensure_ascii=False, indent=2)
                lines.append(f"{key}: {value}")
            lines.append("")

        return "\n".join(lines)


class MonitoringMixin:
    """Миксин для добавления функций мониторинга"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._subscriptions: list[dict[str, Any]] = []
        self._monitoring_task: asyncio.Task | None = None
        self._webhook_callbacks: dict[str, Callable] = {}

    async def monitor_lots(
        self,
        filters: LotsFiltersInput | dict[str, Any],
        callback: Callable[[list[LotResult]], None],
        interval: int = 300,
        max_results: int = 50,
        filter_preset: str | None = None,
    ) -> str:
        """
        Мониторинг новых лотов по подписке

        Args:
            filters: Фильтры для мониторинга
            callback: Функция обратного вызова для новых лотов
            interval: Интервал проверки в секундах
            max_results: Максимальное количество результатов за раз
            filter_preset: Имя предустановленного фильтра

        Returns:
            ID подписки для управления
        """
        subscription_id = (
            f"lots_{len(self._subscriptions)}_{int(datetime.now().timestamp())}"
        )

        # Создание подписки
        subscription = {
            "id": subscription_id,
            "type": "lots",
            "filters": filters,
            "callback": callback,
            "interval": interval,
            "max_results": max_results,
            "filter_preset": filter_preset,
            "last_check": datetime.now(),
            "last_results": set(),
            "active": True,
        }

        self._subscriptions.append(subscription)

        # Запуск мониторинга если еще не запущен
        if not self._monitoring_task or self._monitoring_task.done():
            self._monitoring_task = asyncio.create_task(self._monitoring_loop())

        logger.info(f"Created monitoring subscription {subscription_id}")
        return subscription_id

    async def monitor_contracts(
        self,
        filters: ContractFiltersInput | dict[str, Any],
        callback: Callable[[list[ContractResult]], None],
        interval: int = 600,
        max_results: int = 50,
    ) -> str:
        """Мониторинг новых контрактов"""
        subscription_id = (
            f"contracts_{len(self._subscriptions)}_{int(datetime.now().timestamp())}"
        )

        subscription = {
            "id": subscription_id,
            "type": "contracts",
            "filters": filters,
            "callback": callback,
            "interval": interval,
            "max_results": max_results,
            "last_check": datetime.now(),
            "last_results": set(),
            "active": True,
        }

        self._subscriptions.append(subscription)

        if not self._monitoring_task or self._monitoring_task.done():
            self._monitoring_task = asyncio.create_task(self._monitoring_loop())

        return subscription_id

    async def _monitoring_loop(self):
        """Основной цикл мониторинга"""
        logger.info("Started monitoring loop")

        while True:
            try:
                current_time = datetime.now()

                for subscription in self._subscriptions:
                    if not subscription["active"]:
                        continue

                    # Проверка интервала
                    time_since_last = (
                        current_time - subscription["last_check"]
                    ).total_seconds()
                    if time_since_last < subscription["interval"]:
                        continue

                    await self._check_subscription(subscription)
                    subscription["last_check"] = current_time

                # Удаление неактивных подписок
                self._subscriptions = [s for s in self._subscriptions if s["active"]]

                # Если нет активных подписок, завершаем цикл
                if not self._subscriptions:
                    logger.info("No active subscriptions, stopping monitoring loop")
                    break

                await asyncio.sleep(60)  # Проверка каждую минуту

            except asyncio.CancelledError:
                logger.info("Monitoring loop cancelled")
                break
            except Exception as e:
                logger.error(f"Error in monitoring loop: {e}")
                await asyncio.sleep(60)

    async def _check_subscription(self, subscription: dict[str, Any]):
        """Проверка подписки на новые результаты"""
        try:
            subscription_type = subscription["type"]
            filters = subscription["filters"]
            max_results = subscription["max_results"]

            # Получение новых результатов
            if subscription_type == "lots":
                if isinstance(filters, dict):
                    results = await self.search_lots(limit=max_results, **filters)
                else:
                    # Если filters - это объект LotsFiltersInput
                    results = await self.search_lots(
                        limit=max_results
                    )  # Нужна реализация прямой передачи фильтра
            elif subscription_type == "contracts":
                if isinstance(filters, dict):
                    results = await self.search_contracts(limit=max_results, **filters)
                else:
                    results = await self.search_contracts(limit=max_results)
            else:
                logger.error(f"Unknown subscription type: {subscription_type}")
                return

            # Фильтрация новых результатов
            current_ids = {result.id for result in results}
            new_ids = current_ids - subscription["last_results"]

            if new_ids:
                new_results = [r for r in results if r.id in new_ids]
                logger.info(
                    f"Found {len(new_results)} new {subscription_type} for subscription {subscription['id']}"
                )

                # Вызов callback
                try:
                    await subscription["callback"](new_results)
                except Exception as e:
                    logger.error(
                        f"Error in callback for subscription {subscription['id']}: {e}"
                    )

            # Обновление последних результатов
            subscription["last_results"] = current_ids

        except Exception as e:
            logger.error(f"Error checking subscription {subscription['id']}: {e}")

    async def stop_monitoring(self, subscription_id: str) -> bool:
        """Остановка мониторинга по ID подписки"""
        for subscription in self._subscriptions:
            if subscription["id"] == subscription_id:
                subscription["active"] = False
                logger.info(f"Stopped monitoring subscription {subscription_id}")
                return True
        return False

    async def list_subscriptions(self) -> list[dict[str, Any]]:
        """Список активных подписок"""
        return [
            {
                "id": s["id"],
                "type": s["type"],
                "interval": s["interval"],
                "last_check": s["last_check"].isoformat(),
                "active": s["active"],
                "filter_preset": s.get("filter_preset"),
            }
            for s in self._subscriptions
        ]


class BatchingMixin:
    """Миксин для батчинга запросов"""

    async def batch_search_lots(
        self, search_requests: list[dict[str, Any]], max_concurrent: int = 5
    ) -> list[list[LotResult]]:
        """
        Батчевый поиск лотов с ограничением конкурентности

        Args:
            search_requests: Список запросов для поиска
            max_concurrent: Максимальное количество одновременных запросов

        Returns:
            Список результатов для каждого запроса
        """
        semaphore = asyncio.Semaphore(max_concurrent)

        async def search_with_semaphore(request):
            async with semaphore:
                return await self.search_lots(**request)

        tasks = [search_with_semaphore(request) for request in search_requests]
        return await asyncio.gather(*tasks, return_exceptions=True)

    async def batch_search_by_bins(
        self,
        bins: list[str],
        entity_type: str = "lots",
        batch_size: int = 10,
        **search_params,
    ) -> dict[str, list[Any]]:
        """
        Батчевый поиск по списку БИН

        Args:
            bins: Список БИН для поиска
            entity_type: Тип сущности ('lots', 'contracts', 'subjects')
            batch_size: Размер батча
            **search_params: Дополнительные параметры поиска

        Returns:
            Словарь результатов по БИН
        """
        results = {}

        # Разбивка на батчи
        for i in range(0, len(bins), batch_size):
            batch_bins = bins[i : i + batch_size]

            try:
                if entity_type == "lots":
                    batch_results = await self.search_lots(
                        customer_bin=batch_bins, **search_params
                    )
                elif entity_type == "contracts":
                    batch_results = await self.search_contracts(
                        customer_bin=batch_bins, **search_params
                    )
                elif entity_type == "subjects":
                    batch_results = await self.search_subjects(
                        bin_list=batch_bins, **search_params
                    )
                else:
                    raise ValueError(f"Unknown entity type: {entity_type}")

                # Группировка результатов по БИН
                for result in batch_results:
                    if hasattr(result, "customerBin"):
                        bin_key = result.customerBin
                    elif hasattr(result, "bin"):
                        bin_key = result.bin
                    else:
                        continue

                    if bin_key not in results:
                        results[bin_key] = []
                    results[bin_key].append(result)

            except Exception as e:
                logger.error(f"Error in batch search for bins {batch_bins}: {e}")
                continue

        return results


class TelegramMixin:
    """Миксин для интеграции с Telegram ботом"""

    def format_lot_for_telegram(self, lot: LotResult) -> str:
        """Форматирование лота для Telegram"""
        text = f"🔹 **Лот {lot.lotNumber}**\n"

        if lot.nameRu:
            text += f"📝 {lot.nameRu}\n"

        if lot.amount > 0:
            text += f"💰 Сумма: {lot.amount:,.0f} тг\n"

        if lot.customerNameRu:
            text += f"🏢 Заказчик: {lot.customerNameRu}\n"

        if lot.customerBin:
            text += f"🏛️ БИН: {lot.customerBin}\n"

        if lot.tradeMethod:
            text += f"🛒 Способ: {lot.tradeMethod}\n"

        if lot.status:
            text += f"📌 Статус: {lot.status}\n"

        if lot.endDate:
            text += f"⏰ До: {lot.endDate}\n"

        return text

    def format_contract_for_telegram(self, contract: ContractResult) -> str:
        """Форматирование контракта для Telegram"""
        text = f"📋 **Контракт {contract.contractNumber}**\n"

        if contract.supplierNameRu:
            text += f"🏭 Поставщик: {contract.supplierNameRu}\n"

        if contract.customerNameRu:
            text += f"🏢 Заказчик: {contract.customerNameRu}\n"

        if contract.contractSum > 0:
            text += f"💰 Сумма: {contract.contractSum:,.0f} тг\n"

        if contract.signDate:
            text += f"📅 Подписан: {contract.signDate}\n"

        if contract.status:
            text += f"📌 Статус: {contract.status}\n"

        return text

    async def search_lots_for_telegram(
        self, keyword: str, limit: int = 10, **filters
    ) -> str:
        """Поиск лотов с форматированием для Telegram"""
        try:
            lots = await self.search_lots(keyword=keyword, limit=limit, **filters)

            if not lots:
                return f"🔍 По запросу '{keyword}' ничего не найдено"

            response = f"📋 Найдено лотов: {len(lots)}\n\n"

            for lot in lots:
                response += self.format_lot_for_telegram(lot)
                response += "\n" + "─" * 40 + "\n\n"

            return response

        except Exception as e:
            logger.error(f"Error in Telegram search: {e}")
            return f"❌ Ошибка поиска: {str(e)}"


class StatsMixin:
    """Миксин для статистической агрегации"""

    async def get_lots_stats(
        self, results: list[LotResult], group_by: str = "tradeMethod"
    ) -> dict[str, Any]:
        """
        Статистическая агрегация лотов

        Args:
            results: Результаты поиска
            group_by: Поле для группировки

        Returns:
            Статистика по лотам
        """
        if not results:
            return {"total": 0}

        stats = {
            "total": len(results),
            "total_amount": sum(lot.amount for lot in results),
            "avg_amount": sum(lot.amount for lot in results) / len(results),
            "max_amount": max(lot.amount for lot in results),
            "min_amount": min(lot.amount for lot in results if lot.amount > 0),
        }

        # Группировка
        groups = {}
        for lot in results:
            key = getattr(lot, group_by, "Неизвестно")
            if key not in groups:
                groups[key] = {"count": 0, "amount": 0}
            groups[key]["count"] += 1
            groups[key]["amount"] += lot.amount

        stats["groups"] = groups

        return stats


# === ПОЛНЫЙ КЛИЕНТ С МИКСИНАМИ ===


class GoszakupClientFull(
    ExportMixin,
    MonitoringMixin,
    BatchingMixin,
    TelegramMixin,
    StatsMixin,
    GoszakupClient,
):
    """Полная версия клиента со всеми расширениями"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Предустановленные фильтры
        self.filter_presets = {
            "construction_almaty": {
                "keyword": "строительство",
                "regions": ["Алматы"],
                "trade_methods": [TradeMethod.OPEN_TENDER],
                "status": [LotStatus.PUBLISHED, LotStatus.ACCEPTING_APPLICATIONS],
            },
            "it_equipment": {
                "keyword": "компьютер",
                "trade_methods": [
                    TradeMethod.ELECTRONIC_STORE,
                    TradeMethod.REQUEST_QUOTATIONS,
                ],
            },
            "medical_supplies": {
                "keyword": "медицинск",
                "trade_methods": [
                    TradeMethod.OPEN_TENDER,
                    TradeMethod.REQUEST_QUOTATIONS,
                ],
            },
        }

    async def search_with_preset(
        self,
        preset_name: str,
        additional_filters: dict[str, Any] | None = None,
        limit: int = 50,
    ) -> list[LotResult]:
        """
        Поиск с использованием предустановленных фильтров

        Args:
            preset_name: Название предустановки
            additional_filters: Дополнительные фильтры
            limit: Максимальное количество результатов

        Returns:
            Результаты поиска
        """
        if preset_name not in self.filter_presets:
            raise ValueError(f"Unknown filter preset: {preset_name}")

        filters = self.filter_presets[preset_name].copy()

        if additional_filters:
            filters.update(additional_filters)

        return await self.search_lots(limit=limit, **filters)


# Функция для создания процедуры обратного вызова для мониторинга
def create_monitoring_callback(process_func: Callable[[Any], None]) -> Callable:
    """
    Создание callback функции для мониторинга

    Args:
        process_func: Функция для обработки новых результатов

    Returns:
        Асинхронная callback функция
    """

    async def callback(new_results):
        for result in new_results:
            try:
                (
                    await process_func(result)
                    if asyncio.iscoroutinefunction(process_func)
                    else process_func(result)
                )
            except Exception as e:
                logger.error(f"Error processing monitoring result: {e}")

    return callback
